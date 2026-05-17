"""
Build Excel report with 4 sheets: All stocks, Signals, Backtest, Summary stats.
"""

from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scanner.config import REPORTS_DIR
from scanner.utils import fmt_price, logger

# Colors
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def build_excel_report(
    results: pd.DataFrame,
    signals: dict[str, pd.DataFrame],
    scan_date: str | None = None,
    ai_analysis: dict | None = None,
    super_stocks: pd.DataFrame | None = None,
    company_data: dict | None = None,
) -> dict[str, str]:
    """
    Dual mode  → 2 file riêng: report_long_*.xlsx và report_short_*.xlsx
    Single mode→ 1 file:        report_YYYY-MM-DD.xlsx
    Trả về dict {"long": path, "short": path} hoặc {"long": path}.
    """
    if scan_date is None:
        scan_date = date.today().strftime("%Y-%m-%d")

    if "exchange" not in results.columns:
        try:
            from scanner.database import db_cursor
            tickers = results["ticker"].tolist()
            with db_cursor(commit=False) as cur:
                cur.execute(
                    "SELECT ticker, exchange FROM watchlist WHERE ticker = ANY(%s)",
                    (tickers,),
                )
                exch_map = {r["ticker"]: r["exchange"] for r in cur.fetchall()}
            results = results.copy()
            results["exchange"] = results["ticker"].map(exch_map).fillna("")
        except Exception:
            results = results.copy()
            results["exchange"] = ""

    is_dual = "long_buy_signal" in results.columns
    paths: dict[str, str] = {}

    if is_dual:
        paths["long"]  = _save_workbook(
            _build_workbook(results, signals, "long",  scan_date, ai_analysis, super_stocks, company_data),
            REPORTS_DIR / f"report_long_{scan_date}.xlsx",
        )
        paths["short"] = _save_workbook(
            _build_workbook(results, signals, "short", scan_date, ai_analysis=None, super_stocks=None, company_data=company_data),
            REPORTS_DIR / f"report_short_{scan_date}.xlsx",
        )
    else:
        wb = Workbook()
        _sheet_signals(wb, signals, ai_analysis=ai_analysis, overviews=(company_data or {}).get("overviews"))
        if super_stocks is not None and not super_stocks.empty:
            _sheet_super_stocks(wb, super_stocks, scan_date)
        if ai_analysis:
            _sheet_ai(wb, ai_analysis, scan_date)
        _sheet_all(wb, results)
        _sheet_stats(wb, results, scan_date, style=None, company_data=company_data)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        paths["long"] = _save_workbook(wb, REPORTS_DIR / f"report_{scan_date}.xlsx")

    _cleanup_old_reports(keep_days=3)
    return paths


def _cleanup_old_reports(keep_days: int = 3) -> None:
    """Xóa file Excel cũ, chỉ giữ lại keep_days ngày gần nhất cho mỗi style."""
    import re
    pattern = re.compile(r"report(?:_(?:long|short))?_(\d{4}-\d{2}-\d{2})\.xlsx$")
    files: dict[str, list] = {}
    for f in REPORTS_DIR.glob("report*.xlsx"):
        m = pattern.match(f.name)
        if not m:
            continue
        day = m.group(1)
        prefix = f.name.replace(f"_{day}.xlsx", "")
        files.setdefault(prefix, []).append((day, f))

    for prefix, items in files.items():
        items.sort(key=lambda x: x[0], reverse=True)
        for day, f in items[keep_days:]:
            try:
                f.unlink()
                logger.info(f"Xóa report cũ: {f.name}")
            except Exception:
                pass


def _build_workbook(
    results: pd.DataFrame,
    signals: dict[str, pd.DataFrame],
    style: str,
    scan_date: str,
    ai_analysis: dict | None = None,
    super_stocks: pd.DataFrame | None = None,
    company_data: dict | None = None,
) -> "Workbook":
    """Tạo 1 Workbook hoàn chỉnh cho 1 style (long hoặc short)."""
    wb = Workbook()

    overviews    = (company_data or {}).get("overviews")
    shareholders = (company_data or {}).get("shareholders")

    _sheet_signals(wb, signals, ai_analysis=ai_analysis, style_filter=style, overviews=overviews)
    _sheet_nam_giu(wb, results, style=style)
    _sheet_dung_ngoai(wb, results, style=style)
    _sheet_super_stocks(wb, results, scan_date)
    _sheet_rut_chan(wb, style=style, scan_date=scan_date)

    if style == "long" and ai_analysis is not None:
        _sheet_ai(wb, ai_analysis, scan_date)

    _sheet_history(wb, results=results, style=style)

    if shareholders:
        signal_tickers = _extract_signal_tickers(signals, style_filter=style)
        _sheet_co_dong(wb, shareholders, signal_tickers=signal_tickers)

    _sheet_stats(wb, results, scan_date, style=style, company_data=company_data)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _save_workbook(wb: "Workbook", path) -> str:
    """Lưu workbook, fallback timestamp nếu file đang mở."""
    try:
        wb.save(path)
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime("%H%M%S")
        path = path.parent / f"{path.stem}_{ts}{path.suffix}"
        wb.save(path)
        logger.warning(f"File gốc bị khóa, lưu thành: {path.name}")
    logger.info(f"Excel report saved: {path}")
    return str(path)


def _sheet_all(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Tất cả cổ phiếu")

    criteria = ["ema", "vwap", "rsi", "macd", "adx", "obv", "stoch", "candle", "vol"]
    crit_labels = ["EMA", "VWAP", "RSI", "MACD", "ADX", "OBV", "Stoch", "Nến", "Vol"]

    is_dual = "long_buy_signal" in df.columns

    if is_dual:
        headers = [
            "STT", "Mã", "Xu hướng DH", "Xu hướng NH",
            "Tín hiệu DH", "Tín hiệu NH", "Đồng thuận",
            "Lệnh HĐ", "Ngày vào lệnh", "Giá vào lệnh", "Giá hiện tại",
            "Giữ (phiên)", "Lời/Lỗ %",
            "BiasNorm", "Nhận xét", "bScore", "rScore",
        ] + crit_labels
    else:
        headers = [
            "STT", "Mã", "Xu hướng", "Tín hiệu mới",
            "Lệnh HĐ", "Ngày vào lệnh", "Giá vào lệnh", "Giá hiện tại",
            "Giữ (phiên)", "Lời/Lỗ %",
            "BiasNorm", "Nhận xét", "bScore", "rScore",
        ] + crit_labels
    _write_header(ws, headers)

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        trend_str = "↑ TĂNG" if row.get("trend", 0) == 1 else "↓ GIẢM"
        pnl       = row.get("signal_pnl_pct") or row.get("pnl_pct")
        last_sig  = row.get("last_signal_type", "")
        sig_date  = row.get("last_signal_date")
        sig_price = row.get("last_signal_price")
        bars      = row.get("bars_since_signal")

        def _sig_str(buy_col, sell_col):
            if row.get(buy_col):  return "🟢 MUA"
            if row.get(sell_col): return "🔴 BÁN"
            return ""

        if is_dual:
            trend_dh = "↑" if row.get("long_trend",  0) == 1 else "↓"
            trend_nh = "↑" if row.get("short_trend", 0) == 1 else "↓"
            sig_dh   = _sig_str("long_buy_signal",  "long_sell_signal")
            sig_nh   = _sig_str("short_buy_signal", "short_sell_signal")
            both     = "🔥 CẢ 2" if row.get("both_buy") or row.get("both_sell") else ""
            base_row = [
                i - 1,
                row.get("ticker", ""),
                trend_dh, trend_nh,
                sig_dh, sig_nh, both,
                "Nắm giữ" if last_sig == "MUA" else ("Đứng ngoài" if last_sig == "BÁN" else ""),
                str(sig_date)[:10] if sig_date else "",
                round(sig_price, 2) if sig_price else "",
                row.get("close", 0),
                bars if bars is not None else "",
                round(pnl, 2) if pnl is not None else "",
                round(row.get("bias_norm", 0), 1),
                row.get("bias_label", ""),
                row.get("b_score", 0),
                row.get("r_score", 0),
            ]
        else:
            new_sig = _sig_str("buy_signal", "sell_signal")
            base_row = [
                i - 1,
                row.get("ticker", ""),
                trend_str,
                new_sig,
                "Nắm giữ" if last_sig == "MUA" else ("Đứng ngoài" if last_sig == "BÁN" else ""),
                str(sig_date)[:10] if sig_date else "",
                round(sig_price, 2) if sig_price else "",
                row.get("close", 0),
                bars if bars is not None else "",
                round(pnl, 2) if pnl is not None else "",
                round(row.get("bias_norm", 0), 1),
                row.get("bias_label", ""),
                row.get("b_score", 0),
                row.get("r_score", 0),
            ]
        for c in criteria:
            base_row.append("✓" if row.get(f"bull_{c}") else "")

        for j, val in enumerate(base_row, start=1):
            ws.cell(row=i, column=j, value=val)

        bias = row.get("bias_norm", 50)
        fill = None
        if row.get("buy_signal"):
            fill = GREEN_FILL
        elif row.get("sell_signal"):
            fill = RED_FILL
        elif bias >= 70:
            fill = YELLOW_FILL

        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=i, column=j).fill = fill

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


STAR_FILL   = PatternFill("solid", fgColor="FFD700")
STRONG_FILL = PatternFill("solid", fgColor="00B050")


_SS_COLS   = ["ss1", "ss2", "ss3", "ss4", "ss5", "ss6", "ss7"]
_SS_LABELS = [
    "Giá>EMA200", "EMA200 Tăng", "Gần đỉnh 52T",
    "EMA50>EMA200", "Giá>EMA50", "ADX>25", "Vol20>Vol60",
]
_SS_LIQTHRESH = 50e6


def _sheet_super_stocks(wb: Workbook, df: pd.DataFrame, scan_date: str) -> None:
    ws = wb.create_sheet("Siêu cổ phiếu (5-7 tiêu chí)")

    headers = ["STT", "Mã", "Giá", "Score", "Thanh khoản TB 20p (tỷ)"] + _SS_LABELS
    _write_header(ws, headers)

    try:
        from scanner.database import load_avg_turnover
        tk_map = load_avg_turnover(df["ticker"].tolist())
    except Exception:
        tk_map = {}

    rows_data = []
    for _, row in df.iterrows():
        raw_score = row.get("super_score")
        try:
            score = int(raw_score) if raw_score is not None and raw_score == raw_score else 0
        except (ValueError, TypeError):
            score = 0
        ticker = row.get("ticker", "")
        avg_tk = tk_map.get(ticker, 0)
        tk_ty  = avg_tk / 1e6

        if score < 5 or avg_tk < _SS_LIQTHRESH:
            continue
        rows_data.append((avg_tk, score, tk_ty, row))

    rows_data.sort(key=lambda x: x[0], reverse=True)

    for i, (avg_tk, score, tk_ty, row) in enumerate(rows_data, start=2):
        vals = [
            i - 1,
            row.get("ticker", ""),
            row.get("close", ""),
            f"{score}/7",
            round(tk_ty, 1),
        ] + ["✓" if row.get(c) else "✗" for c in _SS_COLS]

        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)

        fill = STAR_FILL if score == 7 else (STRONG_FILL if score >= 6 else GREEN_FILL)
        for j in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=j)
            cell.fill = fill
            if fill == STRONG_FILL:
                cell.font = Font(color="FFFFFF")

        for j, c in enumerate(_SS_COLS, start=6):  # cột 1=STT, shift +1
            ws.cell(row=i, column=j).fill = GREEN_FILL if row.get(c) else RED_FILL

    note_row = len(rows_data) + 3
    ws.cell(row=note_row, column=1, value="Tiêu chí: score>=5/7 và Thanh khoản TB 20 phiên >=50 tỷ VND").font = BOLD
    ws.cell(row=note_row + 1, column=1, value="7/7=Vàng | 6/7=Xanh đậm | 5/7=Xanh nhạt").font = BOLD

    ALMOST_FILL = PatternFill("solid", fgColor="FFF2CC")

    almost_data = []
    for _, row in df.iterrows():
        raw_score = row.get("super_score")
        try:
            score = int(raw_score) if raw_score is not None and raw_score == raw_score else 0
        except (ValueError, TypeError):
            score = 0
        ticker = row.get("ticker", "")
        avg_tk = tk_map.get(ticker, 0)
        if 3 <= score <= 4:
            almost_data.append((avg_tk, score, avg_tk / 1e6, row))

    almost_data.sort(key=lambda x: x[0], reverse=True)
    almost_top5 = almost_data[:5]

    if almost_top5:
        sep_row = note_row + 3
        cell = ws.cell(row=sep_row, column=1, value="⏳ Sắp thành Siêu cổ phiếu (score 3-4/7, top 5 Thanh khoản)")
        cell.font = Font(bold=True, size=11)
        cell.fill = ALMOST_FILL

        hdr_row = sep_row + 1
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=j, value=h)
            c.fill = PatternFill("solid", fgColor="FFE699")
            c.font = Font(bold=True)

        for k, (avg_tk, score, tk_ty, row) in enumerate(almost_top5, start=hdr_row + 1):
            vals = [
                k - hdr_row,
                row.get("ticker", ""),
                row.get("close", ""),
                f"{score}/7",
                round(tk_ty, 1),
            ] + ["✓" if row.get(c) else "✗" for c in _SS_COLS]

            for j, v in enumerate(vals, start=1):
                ws.cell(row=k, column=j, value=v).fill = ALMOST_FILL

            for j, c in enumerate(_SS_COLS, start=6):
                ws.cell(row=k, column=j).fill = GREEN_FILL if row.get(c) else _LIGHT_RED

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def _sheet_ai(wb: Workbook, ai_analysis: dict, scan_date: str) -> None:
    """Sheet phân tích AI: overview + bảng từng tín hiệu."""
    ws = wb.create_sheet("Phân tích AI")

    WRAP = Alignment(wrap_text=True, vertical="top")
    AI_FILL = PatternFill("solid", fgColor="EAF4FB")

    ws.cell(row=1, column=1, value="NHẬN ĐỊNH TỔNG QUAN THỊ TRƯỜNG").font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="2F75B6")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="FFFFFF")
    ws.merge_cells("A1:D1")

    overview = ai_analysis.get("overview", "") or "(AI không phân tích được — kiểm tra GROQ_API_KEY / GEMINI_API_KEY)"
    ws.cell(row=2, column=1, value=overview).alignment = WRAP
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = max(60, len(overview) // 5)

    buy_signals  = ai_analysis.get("buy_signals",  [])
    sell_signals = ai_analysis.get("sell_signals", [])

    row = 4
    for section_label, signals_list in [
        ("PHÂN TÍCH TÍN HIỆU MUA", buy_signals),
        ("PHÂN TÍCH TÍN HIỆU BÁN", sell_signals),
    ]:
        if not signals_list:
            continue
        ws.cell(row=row, column=1, value=section_label).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="375623" if "MUA" in section_label else "9C0006")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        header_row = row
        for j, h in enumerate(["STT", "Mã", "Nhận định AI", "Ngày quét"], start=1):
            c = ws.cell(row=header_row, column=j, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        row += 1

        for stt, item in enumerate(signals_list, start=1):
            ws.cell(row=row, column=1, value=stt)
            ws.cell(row=row, column=2, value=item.get("ticker", "")).font = Font(bold=True)
            cell = ws.cell(row=row, column=3, value=item.get("analysis", ""))
            cell.alignment = WRAP
            cell.fill = AI_FILL
            ws.cell(row=row, column=4, value=scan_date)
            ws.row_dimensions[row].height = 40
            row += 1

        row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 14


def _sheet_signals(
    wb: Workbook,
    signals: dict,
    ai_analysis: dict | None = None,
    style_filter: str | None = None,
    overviews: dict | None = None,
) -> None:
    """Tab đầu tiên: Tín hiệu trong ngày."""
    style_label = {"long": " Dài hạn", "short": " Ngắn hạn"}.get(style_filter or "", "")
    ws = wb.create_sheet(f"Tín hiệu trong ngày{style_label}")

    today_str = date.today().strftime("%d/%m/%Y")

    headers = [
        "STT", "Mã", "Tín hiệu", "Khung",
        "Ngày mua", "Giá mua/bán (ST)",
        "Thanh khoản (tỷ VND)", "BiasNorm",
        "NN Sở hữu %", "NN Room %", "Free Float %",
    ]
    _write_header(ws, headers)

    is_dual = any("long_buy_signal" in df.columns for df in signals.values() if not df.empty)

    all_combos = [
        ("Bứt phá xác nhận", "long_buy_signal",   "long_sell_signal",  "Dài hạn",  "long",  "long_supertrend"),
        ("Đảo chiều giảm",   "long_sell_signal",  "long_buy_signal",   "Dài hạn",  "long",  "long_supertrend"),
        ("Bứt phá xác nhận", "short_buy_signal",  "short_sell_signal", "Ngắn hạn", "short", "short_supertrend"),
        ("Đảo chiều giảm",   "short_sell_signal", "short_buy_signal",  "Ngắn hạn", "short", "short_supertrend"),
    ]
    combos = [c for c in all_combos if style_filter is None or c[4] == style_filter]

    collected: list[tuple] = []
    seen: set[str] = set()

    def _collect(row, signal_label, khung, st_col, fill):
        ticker = row.get("ticker", "")
        key    = f"{ticker}_{signal_label}_{khung}"
        if key in seen:
            return
        seen.add(key)
        tk = float(row.get("turnover") or 0)
        collected.append((tk, key, row, signal_label, khung, st_col, fill))

    if is_dual:
        for signal_label, buy_col, _sell_col, khung, _style, st_col in combos:
            is_buy = "Bứt phá" in signal_label
            src    = signals.get("buy" if is_buy else "sell", pd.DataFrame())
            if src.empty or buy_col not in src.columns:
                continue
            fill = GREEN_FILL if is_buy else RED_FILL
            for _, row in src[src[buy_col].astype(bool)].iterrows():
                _collect(row, signal_label, khung, st_col, fill)
    else:
        for df, label, fill, st_col in [
            (signals.get("buy",  pd.DataFrame()), "Bứt phá xác nhận", GREEN_FILL, "supertrend"),
            (signals.get("sell", pd.DataFrame()), "Đảo chiều giảm",   RED_FILL,   "supertrend"),
        ]:
            for _, row in df.iterrows():
                _collect(row, label, "Dài hạn", st_col, fill)

    collected.sort(key=lambda x: x[0], reverse=True)

    row_idx = 2
    for stt, (tk, _, row, signal_label, khung, st_col, fill) in enumerate(collected, start=1):
        st     = row.get(st_col) or row.get("supertrend") or ""
        ticker = row.get("ticker", "")
        ov     = (overviews or {}).get(ticker, {})
        vals = [
            stt,
            ticker,
            signal_label,
            khung,
            today_str,
            round(float(st), 2) if st else "",
            round(tk / 1e9, 1) if tk else "",
            round(row.get("bias_norm", 0), 1),
            ov.get("foreign_pct",    ""),
            ov.get("foreign_max_pct", ""),
            ov.get("free_float_pct", ""),
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=j, value=v).fill = fill
        row_idx += 1

    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 40

    ws.column_dimensions["J"].width = 60
    _auto_width(ws)
    ws.column_dimensions["J"].width = max(ws.column_dimensions["J"].width, 60)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    note_row = row_idx + 1
    note_cell = ws.cell(row=note_row, column=1,
                        value="* Ưu tiên mã có Thanh khoản cao (≥ 50 tỷ/phiên) để dễ mua/bán khi cần.")
    note_cell.font = Font(italic=True, color="888888")


_TICKER_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_TICKER_ROW_FILLS = [
    PatternFill("solid", fgColor="EBF5FB"),
    PatternFill("solid", fgColor="F5F5F5"),
]


def _extract_signal_tickers(signals: dict, style_filter: str | None = None) -> list[str]:
    seen: list[str] = []
    for df in signals.values():
        if df.empty or "ticker" not in df.columns:
            continue
        for t in df["ticker"].tolist():
            if t not in seen:
                seen.append(t)
    return seen


def _sheet_co_dong(
    wb: Workbook,
    shareholders: dict[str, pd.DataFrame],
    signal_tickers: list[str] | None = None,
) -> None:
    """Sheet cổ đông lớn: signal tickers trước (nổi bật), sau đó các mã còn lại."""
    if not shareholders:
        return

    ws = wb.create_sheet("Cổ đông lớn")
    headers = ["STT", "Mã", "Cổ đông", "Số cổ phần", "% Sở hữu"]
    _write_header(ws, headers)

    sig_set = set(signal_tickers or [])
    ordered = [t for t in (signal_tickers or []) if t in shareholders]
    ordered += [t for t in shareholders if t not in sig_set]

    row_idx = 2
    stt = 1
    for i, ticker in enumerate(ordered):
        df = shareholders.get(ticker, pd.DataFrame())
        if df.empty:
            continue

        is_signal = ticker in sig_set
        hdr_fill  = _TICKER_HDR_FILL if is_signal else HEADER_FILL
        label     = f"{ticker}  ★ Tín hiệu hôm nay" if is_signal else ticker
        for j in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=j).fill = hdr_fill
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, color="FFFFFF", size=11)
        ws.merge_cells(f"A{row_idx}:E{row_idx}")
        row_idx += 1

        fill = _TICKER_ROW_FILLS[i % 2]
        for _, r in df.iterrows():
            pct = float(r.get("share_own_percent") or 0) * 100
            vals = [stt, ticker, r.get("share_holder", ""), int(r.get("quantity") or 0), round(pct, 3)]
            for j, v in enumerate(vals, start=1):
                ws.cell(row=row_idx, column=j, value=v).fill = fill
            row_idx += 1
            stt += 1

    note_row = row_idx + 1
    ws.cell(row=note_row, column=1,
            value="* Chỉ hiển thị cổ đông lớn đã công bố. Nguồn: VCI/vnstock").font = Font(italic=True, color="888888")
    _auto_width(ws)
    ws.freeze_panes = "B2"


def _sheet_nam_giu(wb: Workbook, results: pd.DataFrame, style: str = "long") -> None:
    """Vùng xanh: long_trend=1, ngày/giá mua từ SuperTrend flip gần nhất, sort thanh khoản↓."""
    ws = wb.create_sheet("Vùng xanh (Nắm giữ)")
    headers = ["STT", "Mã", "Ngày Mua", "Giữ lệnh (ngày)", "Giá Mua", "Giá Hiện Tại", "Lời/Lỗ %", "Thanh khoản (tỷ)"]
    _write_header(ws, headers)

    p         = f"{style}_"
    trend_col = f"{p}trend" if f"{p}trend" in results.columns else "trend"
    date_col  = f"{p}last_signal_date"  if f"{p}last_signal_date"  in results.columns else "last_signal_date"
    price_col = f"{p}last_signal_price" if f"{p}last_signal_price" in results.columns else "last_signal_price"

    df = results[results.get(trend_col, results.get("trend", pd.Series(dtype=int))) == 1].copy()
    if "turnover" in df.columns:
        df = df.sort_values("turnover", ascending=False)

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ticker = row.get("ticker", "")
        close  = float(row.get("close") or 0)
        tk     = float(row.get("turnover") or 0)
        bd     = str(row.get(date_col)  or "")[:10]
        buy_p  = float(row.get(price_col) or 0)

        try:
            hold = (date.today() - pd.to_datetime(bd).date()).days if bd else ""
        except Exception:
            hold = ""

        pnl     = round((close - buy_p) / buy_p * 100, 2) if buy_p > 0 and close > 0 else None
        pnl_str = f"{pnl:+.2f}%" if pnl is not None else ""

        vals = [i - 1, ticker, bd, hold, buy_p or "", close or "", pnl_str, round(tk / 1e9, 1) if tk else ""]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)
        fill = GREEN_FILL if (pnl is None or pnl >= 0) else YELLOW_FILL
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = fill

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


_LIGHT_RED   = PatternFill("solid", fgColor="FFD7D7")
_LIGHT_GREEN = PatternFill("solid", fgColor="C6EFCE")


def _sheet_dung_ngoai(wb: Workbook, results: pd.DataFrame, style: str = "long") -> None:
    """Vùng đỏ: long_trend=-1, ngày/giá bán từ SuperTrend flip gần nhất, sort thanh khoản↓."""
    ws = wb.create_sheet("Vùng đỏ (Đứng ngoài)")
    headers = ["STT", "Mã", "Ngày Bán", "Đứng ngoài (ngày)", "Giá Bán", "Giá Hiện Tại", "Tránh lỗ", "Thanh khoản (tỷ)"]
    _write_header(ws, headers)

    p         = f"{style}_"
    trend_col = f"{p}trend" if f"{p}trend" in results.columns else "trend"
    date_col  = f"{p}last_signal_date"  if f"{p}last_signal_date"  in results.columns else "last_signal_date"
    price_col = f"{p}last_signal_price" if f"{p}last_signal_price" in results.columns else "last_signal_price"

    df = results[results.get(trend_col, results.get("trend", pd.Series(dtype=int))) == -1].copy()
    if "turnover" in df.columns:
        df = df.sort_values("turnover", ascending=False)

    pnl_col_idx = 7  # cột "Tránh lỗ" sau khi thêm STT

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ticker = row.get("ticker", "")
        close  = float(row.get("close") or 0)
        tk     = float(row.get("turnover") or 0)
        sd     = str(row.get(date_col)  or "")[:10]
        sell_p = float(row.get(price_col) or 0)

        try:
            hold = (date.today() - pd.to_datetime(sd).date()).days if sd else ""
        except Exception:
            hold = ""

        pnl = round((close - sell_p) / sell_p * 100, 2) if sell_p > 0 and close > 0 else None
        pnl_str = f"{pnl:+.2f}%" if pnl is not None else ""

        vals = [i - 1, ticker, sd, hold, sell_p or "", close or "", pnl_str, round(tk / 1e9, 1) if tk else ""]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)

        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = _LIGHT_RED

        if pnl is not None:
            ws.cell(row=i, column=pnl_col_idx).fill = _LIGHT_GREEN if pnl <= 0 else _LIGHT_RED

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def _sheet_stats(wb: Workbook, df: pd.DataFrame, scan_date: str, style: str | None = None, company_data: dict | None = None) -> None:
    label = {"long": " Dài hạn", "short": " Ngắn hạn"}.get(style or "", "")
    ws = wb.create_sheet(f"Thống kê{label}")

    def write_kv(r, key, val):
        ws.cell(row=r, column=1, value=key).font = BOLD
        ws.cell(row=r, column=2, value=val)

    is_dual = "long_buy_signal" in df.columns
    p = f"{style}_" if style else ""
    if is_dual and style:
        buy_count  = int(df[f"{p}buy_signal"].sum())  if f"{p}buy_signal"  in df.columns else 0
        sell_count = int(df[f"{p}sell_signal"].sum()) if f"{p}sell_signal" in df.columns else 0
        both_count = 0
    elif is_dual:
        buy_count  = int((df["long_buy_signal"]  | df["short_buy_signal"]).sum())
        sell_count = int((df["long_sell_signal"] | df["short_sell_signal"]).sum())
        both_count = int((df.get("both_buy", False) | df.get("both_sell", False)).sum())
    else:
        buy_count  = int(df["buy_signal"].sum())  if "buy_signal"  in df.columns else 0
        sell_count = int(df["sell_signal"].sum()) if "sell_signal" in df.columns else 0
        both_count = 0
    avg_bias = round(df["bias_norm"].mean(), 1) if "bias_norm" in df.columns else 0

    write_kv(1, "Ngày quét", scan_date)
    write_kv(2, "Tổng mã quét", len(df))
    write_kv(3, "Tín hiệu MUA", buy_count)
    write_kv(4, "Tín hiệu BÁN", sell_count)
    if is_dual:
        write_kv(5, "Đồng thuận cả 2", both_count)
    write_kv(6, "BiasNorm trung bình", avg_bias)

    ws.cell(row=7, column=1, value="Phân phối BiasNorm").font = BOLD
    buckets = [(0, 25, "Rất yếu"), (25, 45, "Yếu"), (45, 55, "Trung tính"), (55, 75, "Mạnh"), (75, 100, "Rất mạnh")]
    for r, (lo, hi, lbl) in enumerate(buckets, start=8):
        count = int(((df["bias_norm"] >= lo) & (df["bias_norm"] < hi)).sum())
        ws.cell(row=r, column=1, value=f"{lbl} ({lo}-{hi})")
        ws.cell(row=r, column=2, value=count)

    ws.cell(row=14, column=1, value="Top 5 mạnh nhất (BiasNorm)").font = BOLD
    top5 = df.nlargest(5, "bias_norm")[["ticker", "bias_norm"]]
    for r, (_, tr) in enumerate(top5.iterrows(), start=15):
        ws.cell(row=r, column=1, value=tr["ticker"])
        ws.cell(row=r, column=2, value=round(tr["bias_norm"], 1))

    ws.cell(row=21, column=1, value="Top 10 Mạnh + Thanh khoản cao").font = BOLD
    top10_headers = ["STT", "Mã", "BiasNorm", "Thanh khoản (tỷ)", "Xu hướng DH", "Xu hướng NH"]
    for j, h in enumerate(top10_headers, start=1):
        cell = ws.cell(row=22, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    if "turnover" in df.columns and "bias_norm" in df.columns:
        tmp = df.copy()
        tmp["turnover"] = pd.to_numeric(tmp["turnover"], errors="coerce").fillna(0)
        tmp["bias_norm"] = pd.to_numeric(tmp["bias_norm"], errors="coerce").fillna(0)

        max_tk = tmp["turnover"].max() or 1
        max_bn = tmp["bias_norm"].max() or 1
        tmp["_score"] = (tmp["bias_norm"] / max_bn) * 0.6 + (tmp["turnover"] / max_tk) * 0.4
        top10 = tmp.nlargest(10, "_score")

        for r, (stt, (_, tr)) in enumerate(zip(range(1, 11), top10.iterrows()), start=23):
            tk_ty    = round(tr["turnover"] / 1e9, 1) if tr["turnover"] else ""
            trend_dh = "↑" if tr.get("long_trend", 0) == 1 else "↓"
            trend_nh = "↑" if tr.get("short_trend", 0) == 1 else "↓"
            vals = [stt, tr["ticker"], round(tr["bias_norm"], 1), tk_ty, trend_dh, trend_nh]
            for j, v in enumerate(vals, start=1):
                ws.cell(row=r, column=j, value=v)
            fill = GREEN_FILL if tr.get("long_trend", 0) == 1 else RED_FILL
            for j in range(1, 7):
                ws.cell(row=r, column=j).fill = fill

    # ── Thống kê thị trường (tổng tài khoản NĐT) ─────────────────────────────
    MARKET_HDR_FILL = PatternFill("solid", fgColor="2E75B6")
    MARKET_ROW_FILL = PatternFill("solid", fgColor="DEEAF1")

    mkt_start = 36
    hdr_cell = ws.cell(row=mkt_start, column=1, value="THỐNG KÊ THỊ TRƯỜNG — Tài khoản nhà đầu tư")
    hdr_cell.font = Font(bold=True, color="FFFFFF")
    hdr_cell.fill = MARKET_HDR_FILL
    ws.merge_cells(f"A{mkt_start}:B{mkt_start}")

    market = (company_data or {}).get("market") or {}
    total    = market.get("total",    0)
    domestic = market.get("domestic", 0)
    foreign  = market.get("foreign",  0)
    as_of    = market.get("as_of",    "N/A")
    source   = market.get("source",   "N/A")

    rows_mkt = [
        ("Tổng tài khoản NĐT",      f"{total:,}" if total else "N/A"),
        ("  Trong nước",             f"{domestic:,}" if domestic else "N/A"),
        ("  Nước ngoài",             f"{foreign:,}" if foreign else "N/A"),
        ("Cập nhật đến",             as_of),
        ("Nguồn",                    source),
    ]
    for offset, (k, v) in enumerate(rows_mkt, start=1):
        r = mkt_start + offset
        ws.cell(row=r, column=1, value=k).fill  = MARKET_ROW_FILL
        ws.cell(row=r, column=2, value=v).fill  = MARKET_ROW_FILL

    _auto_width(ws)


OPEN_FILL   = PatternFill("solid", fgColor="DDEBF7")
CLOSED_FILL = PatternFill("solid", fgColor="F2F2F2")


def _sheet_history(wb: Workbook, results: pd.DataFrame, style: str = "long") -> None:
    """Tab lịch sử lệnh: 1 giao dịch gần nhất/mã (MUA→BÁN), sort theo Thanh khoản cao→thấp."""
    ws = wb.create_sheet("Lịch sử lệnh")

    headers = [
        "STT", "Mã", "Ngày Mua", "Giá Mua",
        "Ngày Bán", "Giá Bán",
        "Lời/Lỗ %", "Trạng thái", "Thanh khoản (tỷ)",
    ]
    _write_header(ws, headers)

    tickers   = results["ticker"].tolist() if "ticker" in results.columns else []
    tk_map    = results.set_index("ticker")["turnover"].to_dict() if "turnover" in results.columns else {}
    close_map = results.set_index("ticker")["close"].to_dict()   if "close"    in results.columns else {}

    trade_df = pd.DataFrame()
    try:
        from scanner.database import load_trade_history_from_ohlcv
        trade_df = load_trade_history_from_ohlcv(style=style, tickers=tickers or None)
    except Exception as e:
        logger.warning(f"_sheet_history: load_trade_history_from_ohlcv failed: {e}")

    if trade_df.empty:
        ws.cell(row=2, column=1, value="Không tính được lịch sử lệnh — kiểm tra log để biết lý do")
        return

    trade_df["_tk"] = trade_df["ticker"].map(tk_map).fillna(0)
    trade_df = trade_df.sort_values("_tk", ascending=False).reset_index(drop=True)

    for i, (_, row) in enumerate(trade_df.iterrows(), start=2):
        ticker     = row.get("ticker", "")
        buy_date   = str(row.get("buy_date")  or "")[:10]
        sell_date  = str(row.get("sell_date") or "")[:10] if pd.notna(row.get("sell_date")) else ""
        buy_price  = row.get("buy_price")
        sell_price = row.get("sell_price")
        status     = row.get("status", "")
        tk_ty      = round(float(row.get("_tk") or 0) / 1e9, 1)

        pnl = row.get("pnl_pct")
        if status == "Dang giu":
            cur_close = close_map.get(ticker)
            if cur_close and buy_price and float(buy_price) > 0:
                pnl = round((float(cur_close) - float(buy_price)) / float(buy_price) * 100, 2)

        try:
            pnl = round(float(pnl), 2) if pnl is not None else None
        except Exception:
            pnl = None

        pnl_str    = f"{pnl:+.2f}%" if pnl is not None else ""
        status_str = "Đang giữ" if status == "Dang giu" else "Đã đóng"

        vals = [
            i - 1,
            ticker,
            buy_date,
            round(float(buy_price), 2)  if buy_price  else "",
            sell_date,
            round(float(sell_price), 2) if sell_price else "",
            pnl_str,
            status_str,
            tk_ty or "",
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)

        if status == "Dang giu":
            row_fill = GREEN_FILL if (pnl is None or pnl >= 0) else RED_FILL
            for j in range(1, len(headers) + 1):
                ws.cell(row=i, column=j).fill = row_fill
        else:
            for j in range(1, len(headers) + 1):
                ws.cell(row=i, column=j).fill = CLOSED_FILL
            ws.cell(row=i, column=7).fill = GREEN_FILL if (pnl is not None and pnl >= 0) else RED_FILL

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


_RUT_CHAN_FILL = PatternFill("solid", fgColor="FCE4D6")


def _sheet_rut_chan(wb: Workbook, style: str, scan_date: str) -> None:
    """Tab tín hiệu rút chân: giá, thanh khoản, ngưỡng ST bị phá, % rút chân."""
    try:
        from scanner.database import db_cursor
        with db_cursor(commit=False) as cur:
            cur.execute(
                """
                SELECT
                    s.ticker,
                    s.signal_type,
                    s.signal_price,
                    s.signal_st,
                    o.close           AS final_close,
                    w.avg_turnover_20d AS turnover
                FROM signals s
                LEFT JOIN ohlcv o
                    ON o.ticker = s.ticker AND o.date = %s
                LEFT JOIN watchlist w
                    ON w.ticker = s.ticker
                WHERE s.signal_date = %s AND s.style = %s AND s.is_fakeout = TRUE
                ORDER BY w.avg_turnover_20d DESC NULLS LAST
                """,
                (scan_date, scan_date, style),
            )
            rows = cur.fetchall()
    except Exception as e:
        logger.warning(f"_sheet_rut_chan query failed: {e}")
        return

    if not rows:
        return

    ws = wb.create_sheet("Tín hiệu rút chân")
    headers = [
        "STT", "Mã", "Tín hiệu", "Giá tín hiệu",
        "Ngưỡng ST bị phá", "Giá đóng cửa", "Rút chân %", "Thanh khoản TB (tỷ)",
    ]
    _write_header(ws, headers)

    for i, row in enumerate(rows, start=2):
        sig_p   = float(row["signal_price"] or 0)
        final   = float(row["final_close"]  or 0)
        st_lvl  = float(row["signal_st"]    or 0)
        tk      = float(row["turnover"]     or 0)
        pct     = round((final - sig_p) / sig_p * 100, 2) if sig_p else None
        pct_str = f"{pct:+.2f}%" if pct is not None else ""

        vals = [
            i - 1,
            row["ticker"],
            row["signal_type"],
            round(sig_p, 2)  if sig_p  else "",
            round(st_lvl, 2) if st_lvl else "",
            round(final, 2)  if final  else "",
            pct_str,
            round(tk / 1e6, 1) if tk else "",
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).fill = _RUT_CHAN_FILL

        if pct is not None and abs(pct) >= 2:
            ws.cell(row=i, column=7).fill = RED_FILL  # cột "Rút chân %" = 7 sau khi thêm STT

    _auto_width(ws)
    ws.freeze_panes = "B2"


def _write_header(ws, headers: list) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
BLUE_FILL   = PatternFill("solid", fgColor="DDEBF7")


def _sheet_positions(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet vị thế đang mở — tất cả mã đang nắm giữ (có buy_date, chưa bán)."""
    ws = wb.create_sheet("Vị thế đang mở")

    pos = df[df["buy_date"].notna()].copy() if "buy_date" in df.columns else pd.DataFrame()

    if pos.empty:
        ws.cell(row=1, column=1, value="Không có vị thế đang mở")
        return

    headers = [
        "STT", "Mã", "Xu hướng", "Ngày mua", "Giá mua",
        "Giá hiện tại", "Lời/Lỗ %", "Lời/Lỗ VND",
        "Hỗ trợ (Stop)", "Tránh lỗ tối đa %",
        "Kháng cự (Target)", "Tiềm năng lãi %",
        "BiasNorm", "Nhận xét",
    ]
    _write_header(ws, headers)

    total_pnl_pct = []

    for i, (_, row) in enumerate(pos.iterrows(), start=2):
        close     = row.get("close", 0) or 0
        buy_price = row.get("buy_price") or 0
        support   = row.get("support", 0) or 0
        resist    = row.get("resistance", 0) or 0
        pnl_pct   = row.get("pnl_pct")
        trend_str = "↑ TĂNG" if row.get("trend", 0) == 1 else "↓ GIẢM"

        max_loss_pct = round((buy_price - support) / buy_price * 100, 2) if buy_price and support else None
        upside_pct   = round((resist - close) / close * 100, 2) if close and resist else None
        pnl_vnd      = round((close - buy_price) / buy_price * 100, 2) if buy_price else None

        vals = [
            i - 1,
            row.get("ticker", ""),
            trend_str,
            str(row.get("buy_date", ""))[:10],
            buy_price,
            close,
            pnl_pct,
            pnl_vnd,
            support,
            f"-{max_loss_pct}%" if max_loss_pct else "",
            resist,
            f"+{upside_pct}%" if upside_pct else "",
            round(row.get("bias_norm", 0), 1),
            row.get("bias_label", ""),
        ]

        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)

        if pnl_pct is not None:
            fill = GREEN_FILL if pnl_pct >= 0 else RED_FILL
            total_pnl_pct.append(pnl_pct)
        else:
            fill = BLUE_FILL

        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = fill

    if total_pnl_pct:
        summary_row = len(pos) + 3
        ws.cell(row=summary_row, column=1, value=f"Tổng {len(pos)} vị thế").font = BOLD
        ws.cell(row=summary_row, column=7,
                value=round(sum(total_pnl_pct) / len(total_pnl_pct), 2)).font = BOLD
        ws.cell(row=summary_row - 1, column=7, value="Lời/Lỗ TB %").font = BOLD

    _auto_width(ws)
    ws.freeze_panes = "B2"


def _sheet_style(wb: Workbook, df: pd.DataFrame, style: str, title: str) -> None:
    """Sheet riêng cho 1 style (long hoặc short) trong dual mode."""
    ws = wb.create_sheet(title)
    p = f"{style}_"

    criteria   = ["ema", "vwap", "rsi", "macd", "adx", "obv", "stoch", "candle", "vol"]
    crit_labels = ["EMA", "VWAP", "RSI", "MACD", "ADX", "OBV", "Stoch", "Nến", "Vol"]

    headers = [
        "STT", "Mã", "Xu hướng", "Tín hiệu",
        "Lệnh HĐ", "Ngày vào lệnh", "Giá vào lệnh", "Giá hiện tại",
        "Giữ (phiên)", "Lời/Lỗ %",
        "Thanh khoản (tỷ VND)", "BiasNorm", "Nhận xét", "bScore", "rScore",
    ] + crit_labels
    _write_header(ws, headers)

    buy_col  = f"{p}buy_signal"
    sell_col = f"{p}sell_signal"
    df_sorted = df.copy()
    df_sorted["_r"] = 1
    if buy_col  in df.columns: df_sorted.loc[df[buy_col],  "_r"] = 0
    if sell_col in df.columns: df_sorted.loc[df[sell_col], "_r"] = 0
    df_sorted = df_sorted.sort_values(["_r", "bias_norm"], ascending=[True, False])

    for i, (_, row) in enumerate(df_sorted.iterrows(), start=2):
        trend_val = row.get(f"{p}trend", 0)
        trend_str = "↑ TĂNG" if trend_val == 1 else "↓ GIẢM"
        sig_str   = "🟢 MUA" if row.get(buy_col) else ("🔴 BÁN" if row.get(sell_col) else "")

        last_sig  = row.get(f"{p}last_signal_type", "")
        sig_date  = row.get(f"{p}last_signal_date")
        sig_price = row.get(f"{p}last_signal_price")
        bars      = row.get(f"{p}bars_since_signal")
        pnl       = row.get(f"{p}signal_pnl_pct")

        tk    = row.get("turnover", 0)
        tk_ty = round(tk / 1e9, 1) if tk else ""

        base_row = [
            i - 1,
            row.get("ticker", ""),
            trend_str,
            sig_str,
            "Nắm giữ" if last_sig == "MUA" else ("Đứng ngoài" if last_sig == "BÁN" else ""),
            str(sig_date)[:10] if sig_date else "",
            round(sig_price, 2) if sig_price else "",
            row.get("close", 0),
            bars if bars is not None else "",
            round(pnl, 2) if pnl is not None else "",
            tk_ty,
            round(row.get("bias_norm", 0), 1),
            row.get("bias_label", ""),
            row.get("b_score", 0),
            row.get("r_score", 0),
        ]
        for c in criteria:
            base_row.append("✓" if row.get(f"bull_{c}") else "")

        for j, val in enumerate(base_row, start=1):
            ws.cell(row=i, column=j, value=val)

        if last_sig == "MUA":
            fill = GREEN_FILL
        elif last_sig == "BÁN":
            fill = RED_FILL
        else:
            fill = None

        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=i, column=j).fill = fill

        if row.get(buy_col) or row.get(sell_col):
            NEW_BUY  = PatternFill("solid", fgColor="00B050")
            NEW_SELL = PatternFill("solid", fgColor="FF0000")
            accent = NEW_BUY if row.get(buy_col) else NEW_SELL
            for j in range(1, 5):  # STT + 3 cột đầu nội dung
                ws.cell(row=i, column=j).fill = accent

    _auto_width(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def _sheet_dong_thuan(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet các mã có tín hiệu đồng thuận cả ngắn hạn lẫn dài hạn."""
    ws = wb.create_sheet("🔥 Đồng thuận cả 2")

    both = df[(df.get("both_buy", False) == True) | (df.get("both_sell", False) == True)].copy()
    if both.empty:
        ws.cell(row=1, column=1, value="Không có mã nào đồng thuận cả 2 style hôm nay")
        return

    headers = ["STT", "Mã", "Loại", "Giá HT", "Lời/Lỗ DH%", "Lời/Lỗ NH%", "BiasNorm", "Nhận xét"]
    _write_header(ws, headers)

    for i, (_, row) in enumerate(both.iterrows(), start=2):
        loai = "🟢 MUA" if row.get("both_buy") else "🔴 BÁN"
        vals = [
            i - 1,
            row.get("ticker", ""),
            loai,
            row.get("close", 0),
            round(row.get("long_signal_pnl_pct",  0), 2) if row.get("long_signal_pnl_pct")  else "",
            round(row.get("short_signal_pnl_pct", 0), 2) if row.get("short_signal_pnl_pct") else "",
            round(row.get("bias_norm", 0), 1),
            row.get("bias_label", ""),
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)
        fill = GREEN_FILL if row.get("both_buy") else RED_FILL
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = fill

    _auto_width(ws)


def _auto_width(ws, max_width: int = 30) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 3, max_width)
